# -*- coding: utf-8 -*-
# @Time    : 2019/12/26 19:36
# @Author  : Xu
# @File    : handle_excel.py
# @Software: PyCharm

from openpyxl import load_workbook
from common.handle_log import logger
from common import project_path
import os

class HandleExcel(object):
    """
    封装Excel操作类
    """

    def __init__(self, file_name, sheet_name=None):
        """
        定义filename、sheet_name变量
        :param filename:
        :param sheet_name:
        """

        logger.debug(f"初始化HandleExcel类的{self.__init__.__name__}方法")
        self.excle_dir = project_path.testData_dir
        self.file_dir = os.path.join(self.excle_dir, file_name)
        self.file_name = self.file_dir
        self.sheet_name = sheet_name
        logger.debug(f"参数filename:{self.file_dir},sheet_name:{sheet_name}")

    def read(self):
        """
        从表格中读取内容
        :return: 返回列表
        """
        wb = None
        flag = 0
        try:
            logger.info(f"调用[{self.read.__name__}]方法")
            # 获取Excel文件对象
            wb = load_workbook(self.file_name)
            flag = 1
            logger.info(f"初始化Excel对象:[{wb}]")
            # 获取表单sheet,如果是None则获取被激活的当前sheet
            if self.sheet_name is None:
                logger.debug(f"sheet_name:{self.sheet_name},则执行ws = wb.active")
                ws = wb.active
                logger.debug(f"获取到当前active是{wb.get_sheet_by_name}")
            else:
                logger.debug(f"sheet_name:{self.sheet_name},则执行ws = wb[self.sheet_name]")
                ws = wb[self.sheet_name]

            # 默认获取第一行为表头
            logger.debug(f"默认获取第一行为表头:{ws.min_row}")
            row_min = ws.min_row
            # 从表格中获取表头
            header = tuple(ws.iter_rows(min_row=row_min, max_row=row_min, values_only=True))[0]
            # 从表格中读取内容
            get_content_list = []
            for content in tuple(ws.iter_rows(min_row=row_min + 1, values_only=True)):
                get_content_list.append(dict(zip(header, content)))
            return get_content_list
        except Exception as e:
            logger.error(f"具体异常是{e}")
        finally:
            if flag == 1:
                logger.info(f"关闭Excel对象:{wb}")
                wb.close()


if __name__ == '__main__':
    dir = project_path.testData_dir

    excel = HandleExcel('case_login.xlsx', 'LogIn')
    data = excel.read()
    print(data)
    print(dir)
    test_data = excel.read()
    print(str(test_data[0]["case_id"]))
    print(str(test_data[0]["description"]))
